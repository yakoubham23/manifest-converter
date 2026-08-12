import pandas as pd
import openpyxl
import shutil
import os

# ============================================================
# TEMPLATE FILES (dans le même dossier que ce script)
# ============================================================
TEMPLATE_PASSENGER = os.path.join(os.path.dirname(__file__), "template_passenger.xlsx")
TEMPLATE_VEHICLE   = os.path.join(os.path.dirname(__file__), "template_vehicle.xlsx")



def _insert_header(ws, text, logo_path, max_col):
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font, Alignment
    
    ws.insert_rows(1)
    ws.row_dimensions[1].height = 55
    
    merge_range = f"B1:{openpyxl.utils.get_column_letter(max_col-1)}1"
    ws.merge_cells(merge_range)
    
    cell = ws.cell(row=1, column=2)
    cell.value = text
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    import os
    if os.path.exists(logo_path):
        # Image gauche
        img_left = OpenpyxlImage(logo_path)
        img_left.height = 65
        img_left.width = int(img_left.width * (65 / img_left.height))
        ws.add_image(img_left, "A1")
        
        # Image droite (alignée avec un peu d'espace via la taille de la colonne)
        img_right = OpenpyxlImage(logo_path)
        img_right.height = 65
        img_right.width = int(img_right.width * (65 / img_right.height))
        last_col = openpyxl.utils.get_column_letter(max_col)
        ws.column_dimensions[last_col].width = 15  # Élargir un peu la dernière colonne pour que le logo ne déborde pas
        ws.column_dimensions["A"].width = 15       # Pareil pour la première
        ws.add_image(img_right, f"{last_col}1")



def _autofit_columns(ws):
    import openpyxl
    for col in ws.columns:
        max_length = 0
        if not col: continue
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row == 1:
                continue
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_length = max(max_length, len(line))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.15
        adjusted_width = min(max(adjusted_width, 12), 45)
        
        current_width = ws.column_dimensions[col_letter].width
        if current_width and current_width > adjusted_width:
            adjusted_width = current_width
            
        ws.column_dimensions[col_letter].width = adjusted_width

def _clean_val(val):
    """Retourne une chaine vide si la valeur est NaN, sinon la valeur."""
    if pd.isna(val):
        return ''
    return val


def process_passenger(passenger_file, output_file, vessel_name='', company_name='', date_dep='', filter_checked_in=False):
    shutil.copy(TEMPLATE_PASSENGER, output_file)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)
        
    headers_pass = ['N°', 'N° passeport', 'Nom', 'Prénom', 'Sexe', 'Date de naissance', 'Nationalité', 'Lieu de départ', "Lieu d'arrivée", 'N° billet']
    for c, h in enumerate(headers_pass, 1):
        ws.cell(row=2, column=c).value = h
        
    for _ in range(16 - len(headers_pass)):
        ws.delete_cols(len(headers_pass) + 1)

    df = pd.read_excel(passenger_file)
    
    counter = 1
    for _, row in df.iterrows():
        passport    = _clean_val(row.get('Travel Document ID', ''))
        nom         = _clean_val(row.get('Surname', ''))
        prenom      = _clean_val(row.get('First Name', ''))

        gender = _clean_val(row.get('Gender', ''))
        if str(gender).upper().startswith('M'):
            gender = 'M'
        elif str(gender).upper().startswith('F'):
            gender = 'F'
        else:
            gender = ''

        dob = row.get('Date of Birth', '')
        if pd.notna(dob):
            try:
                dob = pd.to_datetime(dob).strftime('%d/%m/%Y')
            except Exception:
                dob = str(dob)
        else:
            dob = ''

        nationality = _clean_val(row.get('Nationality', ''))
        from_port   = _clean_val(row.get('From Port UN/LOCODE', ''))
        to_port     = _clean_val(row.get('To Port UN/LOCODE', ''))
        billet      = _clean_val(row.get('Booking Code', ''))
        
        checked_in_raw = str(row.get('Checked-In', '')).strip().upper()
        is_checked_in = checked_in_raw in ('TRUE', 'YES', 'Y', '1')
        
        if filter_checked_in and not is_checked_in:
            continue

        ws.append([
            counter,
            passport,    
            nom,         
            prenom,      
            gender,      
            dob,         
            nationality, 
            from_port,   
            to_port,     
            billet,      
        ])
        counter += 1

    ws.delete_rows(3)
    ws.delete_rows(1)

    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    route = "N/A"
    date_str = "N/A"
    if not df.empty:
        f_port = _clean_val(df.iloc[0].get('From Port UN/LOCODE', ''))
        t_port = _clean_val(df.iloc[0].get('To Port UN/LOCODE', ''))
        if f_port and t_port:
            route = f"{f_port}/{t_port}"
        
        dep_time = df.iloc[0].get('Departure Time', None)
        if pd.notna(dep_time):
            try:
                date_str = pd.to_datetime(dep_time).strftime('%d-%m-%Y')
            except:
                pass
    
    if date_dep.strip():
        date_str = date_dep.strip()
    
    header_text = f"MANIFEST PASSENGERS {route} {date_str} /{company_name.upper()} /{vessel_name.upper()}"
    import os
    logo_path = os.path.join(os.path.dirname(__file__), "image.png")
    _insert_header(ws, header_text, logo_path, len(headers_pass))

    _autofit_columns(ws)
    wb.save(output_file)
    return output_file


def process_vehicle(vehicle_file, output_file, vessel_name='', company_name='', date_dep='', port_dep='', port_arr=''):
    import openpyxl
    import shutil
    import pandas as pd
    
    shutil.copy(TEMPLATE_VEHICLE, output_file)
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row - 3)
        
    headers_veh = ['N°', 'Nom', 'Prénom', 'N° Véhicule', 'Marque', 'Modèle de véhicule', 'Genre', 'Lieu de départ', "Lieu d'arrivée", 'N° billet']
    for c, h in enumerate(headers_veh, 1):
        ws.cell(row=2, column=c).value = h
        
    for _ in range(16 - len(headers_veh)):
        ws.delete_cols(len(headers_veh) + 1)

    with open(vehicle_file, 'rb') as f:
        magic = f.read(10)

    is_html_xls = magic.startswith(b'<html') or magic.startswith(b'<HTML')

    counter = 1
    if is_html_xls:
        dfs = pd.read_html(vehicle_file, encoding='utf-8', header=None)
        df = dfs[0]

        def clean(val):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return ''
            return str(val).replace('\xa0', ' ').replace('\u00a0', ' ').strip()

        def find_col(header_row, label):
            for i, v in enumerate(header_row):
                if label in clean(v):
                    return i
            return None

        header_row_idx = None
        for ri in range(0, min(15, len(df))):
            row_vals = df.iloc[ri].tolist()
            if any('VMA07' in clean(v) for v in row_vals):
                header_row_idx = ri
                break

        if header_row_idx is None:
            raise ValueError("Impossible de trouver les en-têtes VMA dans le fichier.")

        hrow = df.iloc[header_row_idx].tolist()
        col_booking = find_col(hrow, 'VMA06')
        col_cat    = find_col(hrow, 'VMA07')   
        col_make   = find_col(hrow, 'VMA09')   
        col_model  = find_col(hrow, 'VMA10')   
        col_reg    = find_col(hrow, 'VMA11')   
        col_name   = find_col(hrow, 'VMA12')   

        vehicles_grouped = {}
        for ri in range(header_row_idx + 2, len(df)):
            row = df.iloc[ri]

            row_str = ' '.join(clean(v) for v in row.tolist())
            if 'VMA06' in row_str or 'VMA07' in row_str or 'VMA13' in row_str                or 'VMA04' in row_str or 'VMA15' in row_str or 'VMA16' in row_str                or 'VMA19' in row_str:
                continue

            booking_id = clean(row.iloc[col_booking]) if col_booking is not None else ''
            if not booking_id: booking_id = clean(row.iloc[3])
            cat_code  = clean(row.iloc[col_cat])   if col_cat  is not None else ''
            make      = clean(row.iloc[col_make])  if col_make is not None else ''
            model     = clean(row.iloc[col_model]) if col_model is not None else ''
            reg_num   = clean(row.iloc[col_reg])   if col_reg  is not None else ''
            full_name = clean(row.iloc[col_name])  if col_name is not None else ''

            if reg_num == '' and make == '' and full_name == '':
                continue

            parts       = full_name.rsplit(' ', 1)
            nom_prop    = parts[0].strip() if len(parts) > 0 else ''
            prenom_prop = parts[1].strip() if len(parts) > 1 else ''

            cat_upper = cat_code.upper()
            if cat_upper in ('TRA1', 'TRA2', 'REM3'):
                vhl_rmq = 'RMQ'
            elif cat_upper in ('BIKE', 'MOTO', 'MOTOC', 'MOTOB', 'MOTOS'):
                vhl_rmq = 'BIKE'
            else:
                vhl_rmq = 'VHL'
                
            veh_data = {
                'vhl_rmq': vhl_rmq,
                'nom_prop': nom_prop,
                'prenom_prop': prenom_prop,
                'reg_num': reg_num,
                'make': make,
                'model': model,
                'booking_id': booking_id
            }
            if booking_id not in vehicles_grouped:
                vehicles_grouped[booking_id] = []
            vehicles_grouped[booking_id].append(veh_data)

        # Output grouped vehicles
        for b_id, vehs in vehicles_grouped.items():
            def sort_key(v):
                if v['vhl_rmq'] == 'VHL': return 0
                if v['vhl_rmq'] == 'BIKE': return 1
                return 2
            
            sorted_vehs = sorted(vehs, key=sort_key)
            parent_reg_num = ''
            if sorted_vehs:
                parent_reg_num = sorted_vehs[0]['reg_num']
                
            for v in sorted_vehs:
                if v['vhl_rmq'] == 'RMQ' and parent_reg_num:
                    v['reg_num'] = parent_reg_num

                ws.append([
                    counter,
                    v['nom_prop'],    
                    v['prenom_prop'], 
                    v['reg_num'],       
                    v['make'],        
                    v['model'],       
                    v['vhl_rmq'],     
                    port_dep,   
                    port_arr,
                    v['booking_id']
                ])
                counter += 1

    else:
        df = pd.read_excel(vehicle_file)
        
        vehicles_grouped = {}
        for _, row in df.iterrows():
            nom_prop    = _clean_val(row.get('First Primary Driver Last Name', ''))
            prenom_prop = _clean_val(row.get('First Primary Driver First Name', ''))
            veh_num     = _clean_val(row.get('First Primary Vehicle Registration Number', ''))
            chassis     = _clean_val(row.get('First Primary Vehicle VIN Number', ''))
            marque      = _clean_val(row.get('Category Code', ''))
            modele      = _clean_val(row.get('Category Name', ''))
            booking_id  = _clean_val(row.get('Booking Code', ''))

            cat_upper = str(marque).strip().upper()
            if cat_upper in ('TRA1', 'TRA2', 'REM3'):
                vhl_rmq = 'RMQ'
            elif cat_upper in ('BIKE', 'MOTO', 'MOTOC', 'MOTOB', 'MOTOS'):
                vhl_rmq = 'BIKE'
            else:
                vhl_rmq = 'VHL'
                
            group_key = f"{nom_prop}_{prenom_prop}"
            veh_data = {
                'vhl_rmq': vhl_rmq,
                'nom_prop': nom_prop,
                'prenom_prop': prenom_prop,
                'reg_num': veh_num,
                'chassis': chassis,
                'make': marque,
                'model': modele,
                'booking_id': booking_id
            }
            if group_key not in vehicles_grouped:
                vehicles_grouped[group_key] = []
            vehicles_grouped[group_key].append(veh_data)
            
        for g_id, vehs in vehicles_grouped.items():
            def sort_key(v):
                if v['vhl_rmq'] == 'VHL': return 0
                if v['vhl_rmq'] == 'BIKE': return 1
                return 2
            
            sorted_vehs = sorted(vehs, key=sort_key)
            parent_reg_num = ''
            if sorted_vehs:
                parent_reg_num = sorted_vehs[0]['reg_num']
                
            for v in sorted_vehs:
                if v['vhl_rmq'] == 'RMQ' and parent_reg_num:
                    v['reg_num'] = parent_reg_num

                ws.append([
                    counter,
                    v['nom_prop'],    
                    v['prenom_prop'], 
                    v['reg_num'],     
                    v['make'],        
                    v['model'],       
                    v['vhl_rmq'],     
                    port_dep,
                    port_arr,
                    v['booking_id']
                ])
                counter += 1

    ws.delete_rows(3)
    ws.delete_rows(1)

    from openpyxl.styles import PatternFill, Font
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    route = "N/A"
    date_str = "N/A"
    vessel = vessel_name.upper()
    
    if is_html_xls:
        for r_idx in range(min(15, len(df))):
            row_vals = [str(x) for x in df.iloc[r_idx].tolist()]
            for val in row_vals:
                if 'VMA01:' in val:
                    parts = val.split('Route', 1)
                    if len(parts) > 1:
                        raw_route = parts[1].replace('\xa0', '').strip()
                        if 'Algiers - Alicante' in raw_route: route = 'ALG/ALC'
                        elif 'Alicante - Algiers' in raw_route: route = 'ALC/ALG'
                        elif 'Oran - Alicante' in raw_route: route = 'ORN/ALC'
                        elif 'Alicante - Oran' in raw_route: route = 'ALC/ORN'
                        else: route = raw_route
                elif 'VMA02:' in val:
                    import re
                    m = re.search(r'(\d{2}-\d{2}-\d{4})', val)
                    if m:
                        date_str = m.group(1)
                elif 'VMA03:' in val and not vessel:
                    parts = val.split('Vessel', 1)
                    if len(parts) > 1:
                        vessel = parts[1].replace('\xa0', '').strip().upper()
                        
    if date_dep.strip():
        date_str = date_dep.strip()
                        
    header_text = f"MANIFEST VEHICLES {route} {date_str} /{company_name.upper()} /{vessel}"
    import os
    logo_path = os.path.join(os.path.dirname(__file__), "image.png")
    _insert_header(ws, header_text, logo_path, len(headers_veh))

    _autofit_columns(ws)
    wb.save(output_file)
    return output_file
