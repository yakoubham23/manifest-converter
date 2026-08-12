import streamlit as st
import os
import tempfile
from converter import process_passenger, process_vehicle
from converter_gema import process_gema_passenger, process_gema_vehicle, process_gema_full

st.set_page_config(page_title="Convertisseur PSC APIS", layout="wide")

st.title("🚢 Convertisseur Excel PSC APIS")
st.markdown("Convertit les manifestes de passagers et de véhicules au format PSC APIS demandé.")

format_mode = st.radio("Sélectionnez le format de destination :", ["Format Standard (PAF)", "Format GEMA"], horizontal=True)

if format_mode == "Format GEMA":
    st.header("Conversion GEMA (Passagers & Véhicules)")
    st.markdown("Pour le format GEMA, les deux manifestes sont générés dans un **seul et unique fichier Excel**.")
    st.info("Veuillez importer les fichiers bruts d'origine (ex: ALGMAR...xlsx et getjobid...xls).")
    
    col_p, col_v = st.columns(2)
    with col_p:
        passenger_file = st.file_uploader("1. Importer le Manifeste des Passagers (.xlsx)", type=["xlsx"], key="gema_pass_f")
        filter_present_y = st.checkbox("Filtre Présenté Uniquement 'Y'", value=False, key="filt_p_gema")
    with col_v:
        vehicle_file = st.file_uploader("2. Importer le Manifeste des Véhicules (.xlsx ou .xls)", type=["xlsx", "xls"], key="gema_veh_f")

    if st.button("Générer le fichier unique GEMA", key="btn_gema"):
        if not passenger_file or not vehicle_file:
            st.warning("Veuillez importer à la fois le manifeste passagers ET le manifeste véhicules pour générer le fichier GEMA complet.")
        else:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp1:
                tmp1.write(passenger_file.getvalue())
                pass_path = tmp1.name
                
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp2:
                tmp2.write(vehicle_file.getvalue())
                veh_path = tmp2.name

            output_path = os.path.join(tempfile.gettempdir(), "GEMA_Complet.xlsx")

            try:
                with st.spinner("Conversion GEMA en cours..."):
                    process_gema_full(pass_path, veh_path, output_path, filter_present_y=filter_present_y)
                        
                st.success("✅ Fichier GEMA complet généré avec succès !")

                with open(output_path, "rb") as f:
                    st.download_button(
                        label="📥 Télécharger le Fichier GEMA Complet",
                        data=f,
                        file_name="GEMA_Complet.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_gema"
                    )
            except Exception as e:
                st.error(f"Erreur lors de la conversion GEMA : {e}")
            finally:
                if os.path.exists(pass_path):
                    os.remove(pass_path)
                if os.path.exists(veh_path):
                    os.remove(veh_path)

else:
    # Format Standard (PAF)
    tab1, tab2 = st.tabs(["👥 Passagers", "🚗 Véhicules"])

    with tab1:
        st.header("Conversion du Manifeste Passagers (PAF)")
        passenger_file = st.file_uploader("Importer le Manifeste des Passagers (.xlsx)", type=["xlsx"], key="pass_f")
        filter_checked_in = st.checkbox("Ne gnrer que les passagers enregistrs (Checked-In = True)", value=False, key="filt_paf_p")
        col1, col2, col3 = st.columns(3)
        with col1:
            comp_pass = st.text_input("Compagnie", value="NOURIS EL BAHR FERRIES", key="comp_p")
        with col2:
            vess_pass = st.text_input("Navire", value="CRACOVIA", key="vess_p")
        with col3:
            date_pass = st.text_input("Date de départ", value="", placeholder="ex: 12-07-2026", key="date_p")


        if st.button("Générer le fichier Passagers", key="btn_pass"):
            if not passenger_file:
                st.warning("Veuillez importer le manifeste des passagers.")
            else:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(passenger_file.getvalue())
                    pass_path = tmp.name

                output_path = os.path.join(tempfile.gettempdir(), "PSC_APIS_Passagers.xlsx")

                try:
                    with st.spinner("Conversion en cours..."):
                        process_passenger(pass_path, output_path, vessel_name=vess_pass, company_name=comp_pass, date_dep=date_pass, filter_checked_in=filter_checked_in)
                    st.success(f"✅ Conversion réussie ! {passenger_file.name} converti au {format_mode}.")

                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="📥 Télécharger le fichier Passagers",
                            data=f,
                            file_name="PSC_APIS_Passagers.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_p"
                        )
                except Exception as e:
                    st.error(f"Erreur lors de la conversion : {e}")
                finally:
                    if os.path.exists(pass_path):
                        os.remove(pass_path)

    with tab2:
        st.header("Conversion du Manifeste Véhicules (PAF)")
        vehicle_file = st.file_uploader("Importer le Manifeste des Véhicules (.xlsx ou .xls)", type=["xlsx", "xls"], key="veh_f")
        col1, col2, col3 = st.columns(3)
        with col1:
            comp_veh = st.text_input("Compagnie", value="NOURIS EL BAHR FERRIES", key="comp_v")
        with col2:
            vess_veh = st.text_input("Navire", value="CRACOVIA", key="vess_v")
        with col3:
            date_veh = st.text_input("Date de départ", value="", placeholder="ex: 12-07-2026", key="date_v")

        col4, col5 = st.columns(2)
        with col4:
            port_dep_veh = st.text_input("Lieu de départ (ex: ALC)", key="port_dep_v")
        with col5:
            port_arr_veh = st.text_input("Lieu d'arrivée (ex: ALG)", key="port_arr_v")


        if st.button("Générer le fichier Véhicules", key="btn_veh"):
            if not vehicle_file:
                st.warning("Veuillez importer le manifeste des véhicules.")
            else:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(vehicle_file.getvalue())
                    veh_path = tmp.name

                output_path = os.path.join(tempfile.gettempdir(), "PSC_APIS_Vehicules.xlsx")

                try:
                    with st.spinner("Conversion en cours..."):
                        process_vehicle(veh_path, output_path, vessel_name=vess_veh, company_name=comp_veh, date_dep=date_veh, port_dep=port_dep_veh, port_arr=port_arr_veh)
                            
                    st.success(f"✅ Conversion réussie ! {vehicle_file.name} converti au {format_mode}.")

                    import openpyxl
                    wb_stat = openpyxl.load_workbook(output_path)
                    ws_stat = wb_stat.active
                    vhl_count = rmq_count = bike_count = 0
                    for r in range(4, ws_stat.max_row + 1):
                        v02 = ws_stat.cell(r, 2).value
                        if v02 == 'VHL':  vhl_count  += 1
                        elif v02 == 'RMQ': rmq_count += 1
                        elif v02 == 'BIKE': bike_count += 1
                    total = vhl_count + rmq_count + bike_count

                    st.markdown("---")
                    st.markdown("### 📊 Statistiques de conversion")
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Total", total)
                    col2.metric("🚗 Voitures (VHL)", vhl_count)
                    col3.metric("🚛 Remorques (RMQ)", rmq_count)
                    col4.metric("🏍️ Motos (BIKE)", bike_count)

                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="📥 Télécharger le fichier Véhicules",
                            data=f,
                            file_name="PSC_APIS_Vehicules.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_v"
                        )
                except Exception as e:
                    st.error(f"Erreur lors de la conversion : {e}")
                finally:
                    if os.path.exists(veh_path):
                        os.remove(veh_path)
