import pandas as pd
import openpyxl
import os
import re

# GEMA template
GEMA_TEMPLATE = os.path.join(os.path.dirname(__file__), "ExcelTemplate_PscApis_SEA.xlsx")

def _clean_val(val):
    if pd.isna(val): return ''
    s = str(val).replace('\xa0', ' ').strip()
    return s if s.lower() != 'nan' else ''

def _autofit_columns_gema(ws):
    for col in ws.columns:
        max_length = 0
        if not col: continue
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        
        for cell in col:
            # Skip rows 1, 2, 3 (GEMA headers)
            if cell.row <= 3:
                continue
            try:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_length = max(max_length, len(line))
            except:
                pass
        
        adjusted_width = max_length + 1.5
        adjusted_width = min(max(adjusted_width, 15), 50)
        
        current_width = ws.column_dimensions[col_letter].width
        if current_width and current_width > adjusted_width:
            adjusted_width = current_width
            
        ws.column_dimensions[col_letter].width = adjusted_width

def _unmerge_all(ws):
    if hasattr(ws, 'merged_cells') and ws.merged_cells:
        for range_ in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(range_))

def process_gema_passenger(passenger_file, output_file, filter_present_y=False):
    wb = openpyxl.load_workbook(GEMA_TEMPLATE)
    ws = wb['Passager']
    _unmerge_all(ws)
    
    # Delete test lines
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row)
        
    df = pd.read_excel(passenger_file, dtype=str)
    
    col_passport    = 'Travel Document ID'
    col_nom         = 'Surname'
    col_prenom      = 'First Name'
    col_gender      = 'Gender'
    col_dob         = 'Date of Birth'
    col_nationality = 'Nationality'
    col_from        = 'From Port UN/LOCODE'
    col_to          = 'To Port UN/LOCODE'
    col_booking     = 'Booking Code'
    
    for i in range(len(df)):
        row = df.iloc[i]
        
        boarded_val = row.get('Checked-In', 'False')
        presente = 'Y' if str(boarded_val).lower() in ['true', '1', 'y', 'yes'] else 'N'
            
        if filter_present_y and presente != 'Y':
            continue
            
        b01 = _clean_val(row.get(col_passport))
        b02 = _clean_val(row.get(col_nom))
        b03 = _clean_val(row.get(col_prenom))
        
        gender = _clean_val(row.get(col_gender)).upper()
        b04 = 'M' if 'M' in gender else ('F' if 'F' in gender else '')
        
        dob = _clean_val(row.get(col_dob))
        if ' ' in dob: dob = dob.split(' ')[0]
        if '-' in dob:
            parts = dob.split('-')
            if len(parts) == 3 and len(parts[0]) == 4: # yyyy-mm-dd
                dob = f"{parts[2]}/{parts[1]}/{parts[0]}"
        b05 = dob
        
        b06 = _clean_val(row.get(col_nationality))
        b07 = _clean_val(row.get(col_from))
        b08 = _clean_val(row.get(col_to))
        b09 = ''
        b10 = _clean_val(row.get(col_booking))
        b11 = ''
        b12 = ''
        b13 = ''
        b14 = ''
        b15 = ''
        b16 = presente
        
        ws.append([b01, b02, b03, b04, b05, b06, b07, b08, b09, b10, b11, b12, b13, b14, b15, b16])
        
        # Force text format
        r_idx = ws.max_row
        for c_idx in range(1, 17):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.number_format = '@'
            cell.data_type = 's'
            
    _autofit_columns_gema(ws)
    
    # Active Passager sheet
    wb.active = wb['Passager']
    wb.save(output_file)
    return output_file


def process_gema_vehicle(vehicle_file, passenger_file, output_file):
    # 1. Load Passport dict from passenger_file
    pass_df = pd.read_excel(passenger_file, dtype=str)
    passport_dict = {}
    for i in range(len(pass_df)):
        row = pass_df.iloc[i]
        b_id = str(row.get('Booking Code', '')).strip()
        ppt = str(row.get('Travel Document ID', '')).strip()
        if ppt.lower() == 'nan': ppt = ''
        if b_id:
            passport_dict[b_id] = ppt
            
    if not passport_dict:
        raise ValueError("Le manifeste passagers fourni ne contient pas de colonne 'Booking Code' ou 'Travel Document ID'. Assurez-vous d'avoir uploadé le fichier RAW d'origine, et non pas le fichier GEMA déjà généré !")

    # 2. Load GEMA template & Code dicts
    wb = openpyxl.load_workbook(GEMA_TEMPLATE)
    
    code_ws = wb['Code']
    make_dict = {}
    model_dict = {}
    
    for i in range(1, code_ws.max_row + 1):
        c1 = str(code_ws.cell(i, 1).value).strip()
        c7 = str(code_ws.cell(i, 7).value).strip()
        c9 = str(code_ws.cell(i, 9).value).strip().lower()
        
        if 'fabricant' in c1.lower() or 'make' in c1.lower() or 'marque' in c1.lower():
            make_dict[c9] = c7
        elif 'modèle' in c1.lower() or 'modele' in c1.lower() or 'model' in c1.lower():
            model_dict[c9] = c7
            
    ws = wb['Véhicule']
    _unmerge_all(ws)
    if ws.max_row >= 4:
        ws.delete_rows(4, ws.max_row)
        
    # Read Vehicle File
    with open(vehicle_file, 'rb') as f:
        magic = f.read(10)
    is_html_xls = magic.startswith(b'<html') or magic.startswith(b'<HTML')
    
    if is_html_xls:
        dfs = pd.read_html(vehicle_file, encoding='utf-8', header=None)
        df = dfs[0]
        
        # Trouver la ligne d'en-tête (VMA06: Booking I)
        header_row_idx = -1
        col_booking = None
        col_cat = None
        col_make = None
        col_model = None
        col_reg = None
        col_name = None
        
        for i in range(min(20, len(df))):
            row_str = ' '.join([_clean_val(x) for x in df.iloc[i].tolist()])
            if 'VMA06' in row_str or 'Booking' in row_str or 'VMA07' in row_str:
                header_row_idx = i
                row_vals = df.iloc[i].tolist()
                for c_idx, val in enumerate(row_vals):
                    v = _clean_val(val).upper()
                    if ('VMA06' in v or 'BOOKING' in v) and col_booking is None: col_booking = c_idx
                    elif ('VMA07' in v or 'CATEGO' in v) and col_cat is None: col_cat = c_idx
                    elif ('VMA09' in v or 'MAKE' in v) and col_make is None: col_make = c_idx
                    elif ('VMA10' in v or 'MODEL' in v) and col_model is None: col_model = c_idx
                    elif ('VMA11' in v or 'REGIST' in v) and col_reg is None: col_reg = c_idx
                    elif ('VMA12' in v or 'NAME' in v) and col_name is None: col_name = c_idx
                break
                
        for ri in range(header_row_idx + 2, len(df)):
            row = df.iloc[ri]
            row_str = ' '.join([_clean_val(v) for v in row.tolist()])
            if 'VMA' in row_str: continue # skip repeating headers
            
            b_id      = _clean_val(row.iloc[col_booking]) if col_booking is not None else ''
            cat_code  = _clean_val(row.iloc[col_cat])     if col_cat is not None else ''
            make      = _clean_val(row.iloc[col_make])    if col_make is not None else ''
            model     = _clean_val(row.iloc[col_model])   if col_model is not None else ''
            reg_num   = _clean_val(row.iloc[col_reg])     if col_reg is not None else ''
            full_name = _clean_val(row.iloc[col_name])    if col_name is not None else ''
            
            if reg_num == '' and full_name == '' and booking_id == '':
                continue
                
            parts = full_name.rsplit(' ', 1)
            nom_prop = parts[0].strip() if len(parts) > 0 else ''
            prenom_prop = parts[1].strip() if len(parts) > 1 else ''
            
            # Passport from booking ID
            v01 = passport_dict.get(b_id, '')
            
            # Type
            cat_upper = cat_code.upper()
            if cat_upper in ('TRA1', 'TRA2', 'REM3'): v02 = 'RMQ'
            elif cat_upper in ('BIKE', 'MOTO', 'MOTOC', 'MOTOB', 'MOTOS'): v02 = 'BIKE'
            else: v02 = 'VHL'
            
            v03 = nom_prop
            v04 = prenom_prop
            v05 = reg_num
            v06 = reg_num # Dupliquer
            
            # Mappings for make and model
            make_lower = make.lower()
            v07 = make_dict.get(make_lower, make) # fallback to text if not found
            
            model_lower = model.lower()
            v08 = model_dict.get(model_lower, model) # fallback to text if not found
            
            v09 = ''
            v10 = ''
            
            ws.append([v01, v02, v03, v04, v05, v06, v07, v08, v09, v10])
            
            r_idx = ws.max_row
            for c_idx in range(1, 11):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.number_format = '@'
                cell.data_type = 's'
    else:
        # Fallback to standard xlsx parser (if needed)
        pass

    _autofit_columns_gema(ws)
    wb.active = wb['Véhicule']
    wb.save(output_file)
    return output_file


def process_gema_full(passenger_file, vehicle_file, output_file, filter_present_y=False):
    # Charge le template GEMA UNE SEULE FOIS pour générer les deux onglets
    wb = openpyxl.load_workbook(GEMA_TEMPLATE)
    
    # ------------------ ONGLET PASSAGERS ------------------
    ws_pass = wb['Passager']
    _unmerge_all(ws_pass)
    if ws_pass.max_row >= 4:
        ws_pass.delete_rows(4, ws_pass.max_row)
        
    pass_df = pd.read_excel(passenger_file, dtype=str)
    
    col_passport    = 'Travel Document ID'
    col_nom         = 'Surname'
    col_prenom      = 'First Name'
    col_gender      = 'Gender'
    col_dob         = 'Date of Birth'
    col_nationality = 'Nationality'
    col_from        = 'From Port UN/LOCODE'
    col_to          = 'To Port UN/LOCODE'
    col_booking     = 'Booking Code'
    
    passport_dict = {}
    
    for i in range(len(pass_df)):
        row = pass_df.iloc[i]
        
        b_id = str(row.get(col_booking, '')).strip()
        ppt = str(row.get(col_passport, '')).strip()
        if ppt.lower() == 'nan': ppt = ''
        if b_id:
            if b_id not in passport_dict:
                passport_dict[b_id] = []
            passport_dict[b_id].append({
                'ppt': ppt,
                'surname': _clean_val(row.get(col_nom)).upper(),
                'first_name': _clean_val(row.get(col_prenom)).upper()
            })
            
        boarded_val = row.get('Checked-In', 'False')
        presente = 'Y' if str(boarded_val).lower() in ['true', '1', 'y', 'yes'] else 'N'
            
        if filter_present_y and presente != 'Y':
            continue
            
            
        b01 = _clean_val(row.get(col_passport))
        b02 = _clean_val(row.get(col_nom))
        b03 = _clean_val(row.get(col_prenom))
        
        gender = _clean_val(row.get(col_gender)).upper()
        b04 = 'M' if 'M' in gender else ('F' if 'F' in gender else '')
        
        dob = _clean_val(row.get(col_dob))
        if ' ' in dob: dob = dob.split(' ')[0]
        if '-' in dob:
            parts = dob.split('-')
            if len(parts) == 3 and len(parts[0]) == 4: # yyyy-mm-dd
                dob = f"{parts[2]}/{parts[1]}/{parts[0]}"
        b05 = dob
        
        b06 = _clean_val(row.get(col_nationality))
        b07 = _clean_val(row.get(col_from))
        b08 = _clean_val(row.get(col_to))
        b09 = ''
        b10 = _clean_val(row.get(col_booking))
        b11 = ''
        b12 = ''
        b13 = ''
        b14 = ''
        b15 = ''
        b16 = presente
        
        ws_pass.append([b01, b02, b03, b04, b05, b06, b07, b08, b09, b10, b11, b12, b13, b14, b15, b16])
        
        # Force text format
        r_idx = ws_pass.max_row
        for c_idx in range(1, 17):
            cell = ws_pass.cell(row=r_idx, column=c_idx)
            cell.number_format = '@'
            cell.data_type = 's'
            
    _autofit_columns_gema(ws_pass)
    
    if not passport_dict:
        raise ValueError("Le manifeste passagers fourni ne contient pas de colonne 'Booking Code' ou 'Travel Document ID'. Assurez-vous d'avoir uploadé le fichier RAW d'origine.")

    # ------------------ ONGLET VEHICULES ------------------
    code_ws = wb['Code']
    make_dict = {}
    model_dict = {}
    
    for i in range(1, code_ws.max_row + 1):
        c1 = str(code_ws.cell(i, 1).value).strip()
        c7 = str(code_ws.cell(i, 7).value).strip()
        c9 = str(code_ws.cell(i, 9).value).strip().lower()
        
        if 'fabricant' in c1.lower() or 'make' in c1.lower() or 'marque' in c1.lower():
            make_dict[c9] = c7
        elif 'modèle' in c1.lower() or 'modele' in c1.lower() or 'model' in c1.lower():
            model_dict[c9] = c7
            
    ws_veh = wb['Véhicule']
    _unmerge_all(ws_veh)
    if ws_veh.max_row >= 4:
        ws_veh.delete_rows(4, ws_veh.max_row)
        
    with open(vehicle_file, 'rb') as f:
        magic = f.read(10)
    is_html_xls = magic.startswith(b'<html') or magic.startswith(b'<HTML')
    
    if is_html_xls:
        dfs = pd.read_html(vehicle_file, encoding='utf-8', header=None)
        df = dfs[0]
        
        header_row_idx = -1
        col_booking = None
        col_cat = None
        col_make = None
        col_model = None
        col_reg = None
        col_name = None
        
        for i in range(min(20, len(df))):
            row_str = ' '.join([_clean_val(x) for x in df.iloc[i].tolist()])
            if 'VMA06' in row_str or 'Booking' in row_str or 'VMA07' in row_str:
                header_row_idx = i
                row_vals = df.iloc[i].tolist()
                for c_idx, val in enumerate(row_vals):
                    v = _clean_val(val).upper()
                    if ('VMA06' in v or 'BOOKING' in v) and col_booking is None: col_booking = c_idx
                    elif ('VMA07' in v or 'CATEGO' in v) and col_cat is None: col_cat = c_idx
                    elif ('VMA09' in v or 'MAKE' in v) and col_make is None: col_make = c_idx
                    elif ('VMA10' in v or 'MODEL' in v) and col_model is None: col_model = c_idx
                    elif ('VMA11' in v or 'REGIST' in v) and col_reg is None: col_reg = c_idx
                    elif ('VMA12' in v or 'NAME' in v) and col_name is None: col_name = c_idx
                break
                
        for ri in range(header_row_idx + 2, len(df)):
            row = df.iloc[ri]
            row_str = ' '.join([_clean_val(v) for v in row.tolist()])
            if 'VMA' in row_str: continue 
            
            b_id      = _clean_val(row.iloc[col_booking]) if col_booking is not None else ''
            cat_code  = _clean_val(row.iloc[col_cat])     if col_cat is not None else ''
            make      = _clean_val(row.iloc[col_make])    if col_make is not None else ''
            model     = _clean_val(row.iloc[col_model])   if col_model is not None else ''
            reg_num   = _clean_val(row.iloc[col_reg])     if col_reg is not None else ''
            full_name = _clean_val(row.iloc[col_name])    if col_name is not None else ''
            
            if reg_num == '' and full_name == '' and booking_id == '':
                continue
                
            parts = full_name.rsplit(' ', 1)
            nom_prop = parts[0].strip() if len(parts) > 0 else ''
            prenom_prop = parts[1].strip() if len(parts) > 1 else ''
            
            v01 = ''
            if b_id in passport_dict:
                passengers = passport_dict[b_id]
                if len(passengers) == 1:
                    v01 = passengers[0]['ppt']
                else:
                    v_clean = ''.join(e for e in full_name.upper() if e.isalnum() or e == ' ').strip()
                    best_ppt = passengers[0]['ppt']
                    found_exact = False
                    for p in passengers:
                        p1 = f"{p['surname']} {p['first_name']}"
                        p2 = f"{p['first_name']} {p['surname']}"
                        if p1 == v_clean or p2 == v_clean or p1 in v_clean or p2 in v_clean or v_clean in p1 or v_clean in p2:
                            best_ppt = p['ppt']
                            found_exact = True
                            break
                    if not found_exact:
                        best_score = 0
                        v_parts = set(v_clean.split())
                        for p in passengers:
                            p_parts = set(p['surname'].split() + p['first_name'].split())
                            score = len(v_parts.intersection(p_parts))
                            if score > best_score:
                                best_score = score
                                best_ppt = p['ppt']
                    v01 = best_ppt
            
            cat_upper = cat_code.upper()
            if cat_upper in ('TRA1', 'TRA2', 'REM3'): v02 = 'RMQ'
            elif cat_upper in ('BIKE', 'MOTO', 'MOTOC', 'MOTOB', 'MOTOS'): v02 = 'BIKE'
            else: v02 = 'VHL'
            
            v03 = nom_prop
            v04 = prenom_prop
            v05 = reg_num
            v06 = reg_num
            
            make_lower = make.lower()
            v07 = make_dict.get(make_lower, make)
            
            # Fuzzy fallback for make
            if v07 == make:
                make_cleaned = make.lower().replace('-', ' ').strip()
                for m_name, m_code in make_dict.items():
                    if make_cleaned in m_name or m_name in make_cleaned:
                        v07 = m_code
                        break
                    if 'mercedes' in make_cleaned and 'mercedes' in m_name:
                        v07 = m_code
                        break
                    if 'vw' in make_cleaned and 'volkswagen' in m_name:
                        v07 = m_code
                        break
            if v07 == make: v07 = make.upper()
            
            model_lower = model.lower()
            v08 = model_dict.get(model_lower, model)
            
            # Fuzzy fallback for model
            if v08 == model:
                mod_cleaned = model.lower().replace('-', ' ').strip()
                
                # Check models under the same manufacturer code first
                found = False
                for m_name, m_code in model_dict.items():
                    if v07 and m_code.startswith(v07):
                        if mod_cleaned == m_name or mod_cleaned in m_name.split():
                            v08 = m_code
                            found = True
                            break
                        if f"classe {mod_cleaned.replace('class', '').strip()}" in m_name:
                            v08 = m_code
                            found = True
                            break
                        if f"serie {mod_cleaned.replace('series', '').strip()}" in m_name:
                            v08 = m_code
                            found = True
                            break
                            
                if not found:
                    for m_name, m_code in model_dict.items():
                        if v07 and m_code.startswith(v07):
                            if mod_cleaned in m_name:
                                v08 = m_code
                                found = True
                                break
                                
                if not found:
                    for m_name, m_code in model_dict.items():
                        if mod_cleaned == m_name or mod_cleaned in m_name.split():
                            v08 = m_code
                            break
                            
            if v08 == model: v08 = model.upper()
            
            v09 = ''
            v10 = ''
            
            ws_veh.append([v01, v02, v03, v04, v05, v06, v07, v08, v09, v10])
            
            r_idx = ws_veh.max_row
            for c_idx in range(1, 11):
                cell = ws_veh.cell(row=r_idx, column=c_idx)
                cell.number_format = '@'
                cell.data_type = 's'
    
    _autofit_columns_gema(ws_veh)
    
    wb.active = wb['Passager']
    wb.save(output_file)
    return output_file
